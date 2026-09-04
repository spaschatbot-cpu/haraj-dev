"""T817 and T818 — every list exports, and every page renders on real rows.

Both criteria are about the same failure and are tested the same way: **from the
registry, not from a list written here**. `apps.console.navigation.PAGES` is
already the single place a screen exists; a test that enumerates screens a second
time is a second list that stops matching, and the page it forgets is precisely
the page nobody checks.

T818 (I6) is meant literally, and the spec says why: v1 shipped a page that
rendered an empty list while every check was green — `php -l` clean, md5
matching, status 302 as usual, because 302 was the login redirect and said
nothing about the page. **A 200 is not a rendering.** So each page here is
rendered against rows that exist, and the assertion is that the rows are *on the
page* — and `test_deleting_a_display_loop_fails_this_file` proves the assertion
can actually fail, by removing a loop and watching the test go red.

T817 (I5) asks for a real `.xlsx` that opens in Excel without a warning. So the
bytes that come back are re-read with `openpyxl` rather than sniffed for a
content type: a CSV served under a spreadsheet MIME type passes a header check
and produces exactly the dialog the criterion forbids.
"""

from __future__ import annotations

import io
from decimal import Decimal

import pytest
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from apps.accounts.models import Company, User
from apps.auctions.models import Auction, AuctionState, Vehicle, VehicleState
from apps.console.exports import XLSX_CONTENT_TYPE
from apps.console.navigation import PAGES
from apps.core import audit as recorder
from apps.money import services as money
from apps.money.models import Invoice, InvoiceState
from apps.odoo.models import InboundMessage, InboundState

pytestmark = pytest.mark.django_db

TEN_K = Decimal("10000.00")

#: The list screens, and the word that must appear in the workbook they produce.
#: Named by url so a screen that loses its export is a failure here rather than
#: an absence nobody notices.
EXPORTING = {
    "console:auctions": "الرقم",
    "console:vehicles": "رقم اللوت",
    "console:customers": "الاسم",
    "console:invoices": "الرقم",
    "console:partner-decisions": "المزاد",
    "console:money-ledger": "العميل",
    "console:audit": "الفعل",
    "console:odoo-inbox": "المصدر",
}

#: What each sidebar page must actually put on the screen, given the rows built
#: below. A page whose entry here is missing fails `test_every_page_is_covered`,
#: so adding a screen and forgetting to prove it renders is not possible quietly.
MUST_RENDER = {
    "console:home": "الرئيسية",
    "console:auctions": "مزاد الرندرة",
    "console:vehicles": "كامري",
    "console:vehicles-import": "استيراد",
    "console:partner-decisions": "كامري",
    "console:customers": "عميل الرندرة",
    "console:invoices": "INV/818/1",
    # The ledger names the bidder the way every screen does — a company
    # bids under the company's name (`accounts.services.display_name`).
    "console:money-ledger": "شركة الرندرة",
    "console:money-health": "صحة المال",
    "console:audit": "console.render_check",
    "console:odoo-inbox": "payment.posted",
    "console:why-no-bid": "ليه ما يقدرش يزايد؟",
}


@pytest.fixture
def viewer(client) -> User:
    """The owner — the only role that can open every page in one pass."""
    user = User.objects.create_user(
        phone="966500000101", full_name="المالك", password="x"
    )
    user.is_staff = True
    user.console_role = "owner"
    user.save(update_fields=["is_staff", "console_role"])
    client.force_login(user)
    return user


@pytest.fixture
def world(db):
    """One row of everything the console lists. Real rows, made the real way."""
    now = timezone.now()

    customer = User.objects.create_user(
        phone="966555556101", full_name="عميل الرندرة", password="x"
    )
    Company.objects.create(user=customer, name="شركة الرندرة")

    auction = Auction.objects.create(
        number=818,
        title="مزاد الرندرة",
        starts_at=now - timezone.timedelta(hours=1),
        ends_at=now + timezone.timedelta(hours=1),
        state=AuctionState.LIVE,
        deposit_required=TEN_K,
    )
    Vehicle.objects.create(
        auction=auction,
        lot_number=1,
        make="تويوتا",
        model="كامري",
        year=2022,
        reserve_price=Decimal("50000.00"),
        state=VehicleState.AWAITING_DECISION,
        owner_company=customer.company,
    )

    money.deposit_insurance(
        user=customer, amount=Decimal("30000.00"), source="cash", reference="pay-818"
    )
    Invoice.objects.create(
        customer=customer,
        number="INV/818/1",
        amount=Decimal("5000.00"),
        state=InvoiceState.OPEN,
        issued_at=now,
    )
    InboundMessage.objects.create(
        source="odoo",
        event="payment.posted",
        delivery_id="D-818",
        subject_ref="PAY-818",
        payload={"amount": "10000.00"},
        raw_body='{"amount": "10000.00"}',
        state=InboundState.PROCESSED,
        note="تمّت",
    )
    recorder.record(
        action="console.render_check",
        entity=auction,
        note="صفّ حقيقي ليُرندَر",
    )
    return customer


def workbook_of(response) -> list[list[str]]:
    """The download, re-read as a workbook. Not sniffed — opened."""
    assert response.status_code == 200
    assert response["Content-Type"] == XLSX_CONTENT_TYPE
    assert "attachment;" in response["Content-Disposition"]
    assert response["Content-Disposition"].endswith('.xlsx"')

    book = load_workbook(io.BytesIO(response.content))
    sheet = book[book.sheetnames[0]]
    rows = [["" if c.value is None else str(c.value) for c in row] for row in sheet.rows]
    book.close()
    return rows


# ---------------------------------------------------------------------------
# T817 / I5 — every list downloads a real workbook
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("url_name,header", sorted(EXPORTING.items()))
def test_every_list_exports_a_real_workbook(client, viewer, world, url_name, header):
    """Opened with openpyxl, not judged by its MIME type.

    A CSV served as a spreadsheet passes a header check and produces exactly the
    "the file format does not match the extension" dialog I5 forbids.
    """
    response = client.get(reverse(url_name), {"export": "xlsx"})

    rows = workbook_of(response)
    assert rows, url_name
    assert header in rows[0], (url_name, rows[0])
    assert len(rows) > 1, f"{url_name} exported headers and no rows"


def test_every_list_screen_offers_the_download(client, viewer, world):
    """The file is reachable from the page, not only from a url somebody knows."""
    for url_name in EXPORTING:
        body = client.get(reverse(url_name)).content.decode()
        assert "تصدير إلى إكسل" in body, url_name
        assert "export=xlsx" in body, url_name


def test_the_export_carries_the_filter_on_the_screen(client, viewer, world):
    """v1 exported the whole table every time, so people copied rows by hand.

    An export that silently ignores the search box is worse than no export,
    because it looks like it worked.
    """
    User.objects.create_user(
        phone="966555556102", full_name="عميل آخر تماماً", password="x"
    )

    everyone = workbook_of(client.get(reverse("console:customers"), {"export": "xlsx"}))
    filtered = workbook_of(
        client.get(reverse("console:customers"), {"q": "عميل الرندرة", "export": "xlsx"})
    )

    assert len(filtered) < len(everyone)
    names = {row[0] for row in filtered[1:]}
    assert names == {"عميل الرندرة"}


def test_the_page_number_is_not_carried_into_the_export(client, viewer, world):
    """Page two of the screen is not page two of the file — the file is all of it."""
    response = client.get(reverse("console:customers"), {"page": "2"})

    assert "page=2" not in response.content.decode().split("export=xlsx")[0][-120:]


def test_one_customers_ledger_exports_too(client, viewer, world):
    """A detail page that carries a list is a list (I5 says every one of them)."""
    rows = workbook_of(
        client.get(reverse("console:money-customer", args=[world.pk]), {"export": "xlsx"})
    )

    assert "المبلغ" in rows[0]
    assert any("30000.00" in cell for row in rows[1:] for cell in row)


def test_an_amount_reaches_the_file_as_its_own_digits(client, viewer, world):
    """Article 3-2 all the way to the workbook.

    Every cell is written as text, so an amount is not re-typed by Excel into a
    float on the way in — the file is reconciled against the ledger, and a
    reconciliation that has to allow for rounding is not one.
    """
    rows = workbook_of(client.get(reverse("console:invoices"), {"export": "xlsx"}))

    amounts = [cell for row in rows[1:] for cell in row if cell == "5000.00"]
    assert amounts, rows


# ---------------------------------------------------------------------------
# T818 / I6 — every page renders on real rows
# ---------------------------------------------------------------------------


def test_every_page_in_the_registry_is_covered_here():
    """Adding a screen and forgetting to prove it renders is not possible quietly."""
    listed = {page.url_name for page in PAGES}

    assert listed == set(MUST_RENDER), listed ^ set(MUST_RENDER)


@pytest.mark.parametrize("url_name,expected", sorted(MUST_RENDER.items()))
def test_every_page_renders_its_rows(client, viewer, world, url_name, expected):
    """Not a 200 — the row, on the page. That distinction is the whole of I6."""
    response = client.get(reverse(url_name))

    assert response.status_code == 200, url_name
    assert expected in response.content.decode(), url_name


def test_the_lists_say_how_many_rows_they_found(client, viewer, world):
    """A count is what makes an empty page visible as empty.

    v1's page rendered nothing and looked like a page with nothing to show. A
    stated count turns "there are no cars" and "the loop is broken" into two
    different screens.
    """
    for url_name in ("console:auctions", "console:customers", "console:invoices"):
        body = client.get(reverse(url_name)).content.decode()
        assert "صفحة 1 من" in body, url_name


def test_deleting_a_display_loop_fails_this_file(client, viewer, world):
    """The assertion can fail. Proven by making it fail, on purpose.

    Without this, `test_every_page_renders_its_rows` is a promise: it passes
    today, and nobody knows whether it would have caught the v1 page that
    rendered an empty list while every check was green. So the real template's
    display loop is emptied and the same rows are rendered through it — and the
    customer who was on the page is not on it any more.

    Rendered from the source string rather than by editing the file on disk:
    production settings cache templates by design (`test.py` inherits `prod.py`
    deliberately), so a file written mid-test would never be re-read and the
    check would pass for the wrong reason — which is the exact failure this test
    exists to rule out.
    """
    import pathlib

    from django.template import engines

    source = pathlib.Path("templates/console/customers.html").read_text(encoding="utf-8")
    context = {"page": User.objects.all(), "q": ""}

    intact = engines["django"].from_string(source).render(context)
    assert "عميل الرندرة" in intact

    broken_source = source.replace(
        "{% for customer in page %}", "{% for customer in nothing %}"
    )
    assert broken_source != source, "the loop this test breaks has been renamed"

    broken = engines["django"].from_string(broken_source).render(context)
    assert "عميل الرندرة" not in broken, "a broken display loop must fail this file"


@pytest.mark.parametrize("url_name", sorted(MUST_RENDER))
def test_no_template_comment_leaks_onto_the_page(client, viewer, world, url_name):
    """`{# #}` is single-line only — a multi-line one renders as literal text.

    That is how two comments ended up visible in the console's topbar while
    every check stayed green: no test read the page for what should not be
    there. All comments are `{% comment %}` now; this test keeps them that way.
    """
    body = client.get(reverse(url_name)).content.decode()

    assert "{#" not in body, url_name
    assert "#}" not in body, url_name
