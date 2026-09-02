"""One reader and one writer for tabular files.

Everything the platform imports or exports goes through :class:`Sheet`. No
controller opens `csv.writer` itself — a CI check refuses it — because in v1
each export grew its own quoting, its own encoding and its own idea of what an
empty cell was, and the file one screen produced could not be re-uploaded to
the screen next to it.

Two decisions worth stating:

**Format is read from the content, never from the file name.** Operators
rename `.xls` to `.csv` and export `.csv` out of Excel as a real workbook; a
router that trusts the extension fails on both. A ZIP magic number says
xlsx, and anything else is decoded as text.

**Every cell is text, in and out.** A round trip must not turn `01` into `1`
or a reserve price into a float. Parsing into typed values is the caller's
job, where the column's meaning is known.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field

from openpyxl import Workbook, load_workbook

#: Local zip magic. An xlsx file is a zip archive; nothing else we accept is.
ZIP_MAGIC = b"PK\x03\x04"

#: Excel on Windows opens a UTF-8 CSV as mojibake unless it finds a BOM, and
#: the people receiving these files use Excel on Windows.
CSV_ENCODING = "utf-8-sig"


class SheetError(Exception):
    """The file could not be read as a table, with an Arabic reason."""


@dataclass
class Sheet:
    headers: list[str]
    rows: list[list[str]] = field(default_factory=list)

    # -- reading ---------------------------------------------------------

    @classmethod
    def read(cls, data: bytes) -> Sheet:
        """Parse bytes into a table, deciding the format by content."""
        if not data:
            raise SheetError("الملف فارغ")
        if data[:4] == ZIP_MAGIC:
            return cls._read_xlsx(data)
        return cls._read_csv(data)

    @classmethod
    def _read_csv(cls, data: bytes) -> Sheet:
        try:
            text = data.decode(CSV_ENCODING)
        except UnicodeDecodeError as exc:
            raise SheetError(
                "تعذّرت قراءة الملف: الترميز ليس UTF-8 وليس ملف إكسل"
            ) from exc

        reader = csv.reader(io.StringIO(text, newline=""))
        table = [[cell.strip() for cell in row] for row in reader]
        return cls._from_table(table)

    @classmethod
    def _read_xlsx(cls, data: bytes) -> Sheet:
        try:
            workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        except Exception as exc:  # openpyxl raises a zoo of exception types
            raise SheetError(f"تعذّرت قراءة ملف إكسل: {exc}") from exc

        worksheet = workbook[workbook.sheetnames[0]]
        table = [[cls._as_text(cell) for cell in row] for row in worksheet.iter_rows()]
        workbook.close()
        return cls._from_table(table)

    @staticmethod
    def _as_text(cell) -> str:
        value = cell.value
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            # Excel stores every number as a double, so a lot number typed as
            # 7 comes back as 7.0. Writing that back out would change the row.
            return str(int(value))
        return str(value).strip()

    @classmethod
    def _from_table(cls, table: list[list[str]]) -> Sheet:
        rows = [row for row in table if any(cell != "" for cell in row)]
        if not rows:
            raise SheetError("الملف لا يحتوي على أي صف")

        headers = [cell.strip() for cell in rows[0]]
        width = len(headers)
        body = [cls._fit(row, width) for row in rows[1:]]
        return cls(headers=headers, rows=body)

    @staticmethod
    def _fit(row: list[str], width: int) -> list[str]:
        """Pad or trim a row to the header width.

        A short row is normal — spreadsheet editors drop trailing empties —
        and refusing it would reject files people can produce by accident.
        """
        return (row + [""] * width)[:width]

    # -- as records ------------------------------------------------------

    def records(self) -> list[dict[str, str]]:
        """Rows as header→value maps, in file order."""
        return [dict(zip(self.headers, row, strict=True)) for row in self.rows]

    # -- writing ---------------------------------------------------------

    def to_csv(self) -> bytes:
        buffer = io.StringIO(newline="")
        writer = csv.writer(buffer, lineterminator="\r\n")
        writer.writerow(self.headers)
        writer.writerows(self.rows)
        return buffer.getvalue().encode(CSV_ENCODING)

    def to_xlsx(self) -> bytes:
        """A real workbook, not a CSV with the wrong extension.

        Cells are written as text so nothing is re-typed on the way out:
        Excel would happily read a VIN as a number in scientific notation and
        hand back a different string than it was given.
        """
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "data"

        worksheet.append(self.headers)
        for row in self.rows:
            worksheet.append(row)

        for cells in worksheet.iter_rows():
            for cell in cells:
                cell.number_format = "@"

        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()
        return buffer.getvalue()
